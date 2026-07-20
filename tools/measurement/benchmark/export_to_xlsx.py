#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export profile/per_sentence_results.csv (+ per_stage_results.csv) - already
produced by the offline join (profiling/join.py) - into a paste-ready Excel
sheet matching the master workbook Chatterbox_Power_Measurements_final.xlsx's
P2P3_Synthesis layout: headers in row 1, columns A-U, one 11-sentence
benchmark pass (REF_start, A1-A3, B1-B4, C1-C2, REF_end) per data block of
rows 2-12.

No synthesis or profiling logic lives here - this only reads the CSVs the
join already wrote and reshapes/derives a few columns for the paste target.
If a benchmark run used --repeats N, per_sentence_results.csv holds N
back-to-back 11-sentence passes (in recorded order); each pass gets its own
sheet (P2P3_Synthesis, P2P3_Synthesis_pass2, ...), all with the same A2:U12
layout so any of them can be pasted individually.

Usage:
    python -m tools.measurement.benchmark.export_to_xlsx                        # profile/latest
    python -m tools.measurement.benchmark.export_to_xlsx --profile-dir profile/run_20260716_120000
    python -m tools.measurement.benchmark.export_to_xlsx --out-dir SOMEWHERE

With no --profile-dir, this defaults to profile/latest (the most recently
completed profiled run - see tools/monitoring/profiling/__init__.py's start_session()); if
that pointer is missing or stale (no per_sentence_results.csv there yet),
it lists the available profile/run_.../ directories and asks which one to
use, rather than failing outright.

Requires openpyxl (pip install openpyxl), imported lazily so a profiling-only
environment without it doesn't crash - just prints how to install it.
"""
import argparse
import csv
import os

COLUMNS = [
    "id", "tag", "words", "phon", "audio_s", "synth_ms", "RTF", "front_ms",
    "acou_ms", "voco_ms", "write_ms", "pmicE_Wh", "synthP_W", "E/s_Wh",
    "ampE_Wh", "ampMean_mW", "ampPk_mW", "peak_C", "throttled", "cpuE_Wh", "cpuP_W",
]
PASS_SIZE = 11  # REF_start, A1-A3, B1-B4, C1-C2, REF_end
STAGES = ["front_end", "acoustic", "vocoder", "write"]


def _to_number(value):
    """csv.DictReader gives strings; "" / missing -> None."""
    if value in (None, ""):
        return None
    return float(value)


def _to_int(value):
    number = _to_number(value)
    return int(number) if number is not None else None


def _round_or_none(value, ndigits):
    return round(value, ndigits) if value is not None else None


def load_per_sentence_rows(profile_dir):
    path = os.path.join(profile_dir, "per_sentence_results.csv")
    if not os.path.exists(path):
        raise SystemExit(
            "[export_to_xlsx] {} not found. Run `python3 do_tts.py --benchmark --profile "
            "--join` first, or `python -m tools.monitoring.profiling.join --profile-dir {}` if that run "
            "already has per_sample.csv/per_sentence.jsonl but was never joined.".format(
                path, profile_dir,
            )
        )
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_per_stage_rows(profile_dir):
    path = os.path.join(profile_dir, "per_stage_results.csv")
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _split_into_passes(rows, block_size):
    """Split rows (in recorded order) into consecutive fixed-size blocks.
    A trailing partial block (interrupted run) is dropped with a warning -
    the target sheet layout requires exactly one full pass per sheet."""
    blocks = [rows[i:i + block_size] for i in range(0, len(rows), block_size)]
    if blocks and len(blocks[-1]) < block_size:
        print("[export_to_xlsx] {} leftover row(s) don't fill a full {}-row block - "
              "dropping them (partial/interrupted run?).".format(len(blocks[-1]), block_size))
        blocks = blocks[:-1]
    return blocks


def _relabel_ref(sentence_id, position):
    if position == 0:
        return "REF_start"
    if position == PASS_SIZE - 1:
        return "REF_end"
    return sentence_id


def build_sheet_row(row, position):
    """Map one per_sentence_results.csv row (+ its position 0..10 within the
    pass) to the 21 A-U column values, computing the derived ones (RTF,
    synthP_W, E/s_Wh, cpuP_W) exactly as specified, so the pasted block is
    self-contained (no formulas)."""
    audio_s = _to_number(row.get("audio_duration_s"))
    synth_ms = _to_number(row.get("total_synth_ms"))
    pmic_wh = _to_number(row.get("energy_wh"))
    cpu_wh = _to_number(row.get("cpu_energy_wh"))
    amp_wh = _to_number(row.get("amp_energy_wh"))
    amp_mean_w = _to_number(row.get("amp_mean_w"))
    amp_peak_w = _to_number(row.get("amp_peak_w"))

    rtf = (synth_ms / 1000.0 / audio_s) if synth_ms is not None and audio_s else None
    synth_p_w = (pmic_wh * 3.6e6 / synth_ms) if pmic_wh is not None and synth_ms else None
    e_per_s_wh = (pmic_wh / audio_s) if pmic_wh is not None and audio_s else None
    cpu_p_w = (cpu_wh * 3.6e6 / synth_ms) if cpu_wh is not None and synth_ms else None

    throttled_raw = row.get("throttled_any")
    throttled = None if throttled_raw in (None, "") else (1 if throttled_raw == "True" else 0)

    return [
        _relabel_ref(row["sentence_id"], position),
        row.get("complexity_tag"),
        _to_int(row.get("word_count")),
        _to_int(row.get("phoneme_count")),
        _round_or_none(audio_s, 3),
        _round_or_none(synth_ms, 0),
        _round_or_none(rtf, 3),
        _round_or_none(_to_number(row.get("front_end_ms")), 0),
        _round_or_none(_to_number(row.get("acoustic_ms")), 0),
        _round_or_none(_to_number(row.get("vocoder_ms")), 0),
        _round_or_none(_to_number(row.get("write_ms")), 0),
        _round_or_none(pmic_wh, 5),
        _round_or_none(synth_p_w, 3),
        _round_or_none(e_per_s_wh, 5),
        _round_or_none(amp_wh, 5),
        _round_or_none(amp_mean_w * 1000.0 if amp_mean_w is not None else None, 3),
        _round_or_none(amp_peak_w * 1000.0 if amp_peak_w is not None else None, 3),
        _round_or_none(_to_number(row.get("peak_temp")), 1),
        throttled,
        _round_or_none(cpu_wh, 5),
        _round_or_none(cpu_p_w, 3),
    ]


def build_stage_sheet_rows(stage_passes):
    """[pass, id, stage, duration_ms, total_energy_wh, cpu_energy_wh,
    mem_energy_wh] for the reference 'per_stage' sheet - not for pasting."""
    rows = []
    for pass_index, stage_rows in enumerate(stage_passes, start=1):
        for position in range(PASS_SIZE):
            block = stage_rows[position * len(STAGES):(position + 1) * len(STAGES)]
            for stage_row in block:
                energy_j = _to_number(stage_row.get("energy_j"))
                rows.append([
                    pass_index,
                    _relabel_ref(stage_row["sentence_id"], position),
                    stage_row["stage"],
                    _round_or_none(_to_number(stage_row.get("duration_ms")), 0),
                    _round_or_none(energy_j / 3600.0 if energy_j is not None else None, 5),
                    _round_or_none(_to_number(stage_row.get("cpu_energy_wh")), 5),
                    _round_or_none(_to_number(stage_row.get("mem_energy_wh")), 5),
                ])
    return rows


def write_workbook(sentence_passes, stage_passes, out_path):
    try:
        import openpyxl
    except ImportError:
        print("[export_to_xlsx] openpyxl not installed - run `pip install openpyxl` and "
              "re-run `python -m tools.measurement.benchmark.export_to_xlsx` to produce the paste-ready sheet. "
              "per_sentence_results.csv / per_stage_results.csv are unaffected.")
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for pass_index, rows in enumerate(sentence_passes, start=1):
        sheet_name = "P2P3_Synthesis" if pass_index == 1 else "P2P3_Synthesis_pass{}".format(pass_index)
        ws = wb.create_sheet(sheet_name)
        ws.append(COLUMNS)
        for position, row in enumerate(rows):
            ws.append(build_sheet_row(row, position))

    ws_stage = wb.create_sheet("per_stage")
    ws_stage.append(["pass", "id", "stage", "duration_ms", "total_energy_wh", "cpu_energy_wh", "mem_energy_wh"])
    for stage_row in build_stage_sheet_rows(stage_passes):
        ws_stage.append(stage_row)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


def export(profile_dir="profile", out_dir=None):
    out_dir = out_dir or os.path.join(profile_dir, "exports")
    sentence_rows = load_per_sentence_rows(profile_dir)
    stage_rows = load_per_stage_rows(profile_dir)

    sentence_passes = _split_into_passes(sentence_rows, PASS_SIZE)
    stage_passes = _split_into_passes(stage_rows, PASS_SIZE * len(STAGES))

    if not sentence_passes:
        print("[export_to_xlsx] no full {}-sentence pass found in {}/per_sentence_results.csv "
              "- nothing to export.".format(PASS_SIZE, profile_dir))
        return None

    out_path = os.path.join(out_dir, "chatterbox_paste.xlsx")
    written = write_workbook(sentence_passes, stage_passes, out_path)
    if written is None:
        return None

    note = " Additional passes are in sheets P2P3_Synthesis_pass2, ... (same A2:U12 layout)." \
        if len(sentence_passes) > 1 else ""
    print("Wrote {} pass(es) to {}".format(len(sentence_passes), written))
    print("Open chatterbox_paste.xlsx, copy A2:U12 from sheet 'P2P3_Synthesis', paste into "
          "Chatterbox_Power_Measurements_final.xlsx, sheet P2P3_Synthesis, at cell A12." + note)
    return written


def _run_has_results(base_dir, run_id):
    return os.path.exists(os.path.join(base_dir, run_id, "per_sentence_results.csv"))


def _latest_pointer_target(base_dir):
    """Resolve profile/latest (symlink or the Windows-without-symlinks
    latest.txt fallback - see profiling._update_latest_pointer(), which
    always writes the plain run_id as the symlink target / file content) to
    a run directory name, or None if there's no usable pointer."""
    link = os.path.join(base_dir, "latest")
    if os.path.islink(link):
        target = os.readlink(link)
        return os.path.basename(target) if target else None
    txt_pointer = link + ".txt"
    if os.path.isfile(txt_pointer):
        with open(txt_pointer, encoding="utf-8") as f:
            run_id = f.read().strip()
        return run_id or None
    return None


def _resolve_profile_dir(base_dir, explicit):
    """--profile-dir if given; otherwise profile/latest if it points at a
    run that's actually been joined; otherwise list the available
    profile/run_.../ directories (most recent first, per
    profiling.list_run_dirs()) and ask which one to export."""
    if explicit:
        return explicit

    run_id = _latest_pointer_target(base_dir)
    if run_id and _run_has_results(base_dir, run_id):
        return os.path.join(base_dir, run_id)

    import tools.monitoring.profiling as profiling
    candidates = [r for r in profiling.list_run_dirs(base_dir) if _run_has_results(base_dir, r)]
    if not candidates:
        raise SystemExit(
            "[export_to_xlsx] no run under {}/ has a per_sentence_results.csv yet - run "
            "`python3 do_tts.py --benchmark --profile --join` first, or `python -m "
            "tools.monitoring.profiling.join --profile-dir <run>` on an existing run.".format(base_dir)
        )

    print("[export_to_xlsx] profile/latest isn't usable (missing, or that run was never "
          "joined) - pick a run to export:")
    for i, candidate in enumerate(candidates, start=1):
        print("  {}) {}{}".format(i, candidate, "  (most recent)" if i == 1 else ""))
    choice = input("Run number [1]: ").strip() or "1"
    try:
        selected = candidates[int(choice) - 1]
    except (ValueError, IndexError):
        raise SystemExit("[export_to_xlsx] invalid selection: {!r}".format(choice))
    return os.path.join(base_dir, selected)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--profile-dir", default=None,
                         help="Directory holding per_sentence_results.csv / "
                              "per_stage_results.csv for one run (default: "
                              "profile/latest; if that isn't usable, prompts "
                              "interactively with the available runs)")
    parser.add_argument("--out-dir", default=None,
                         help="Default: <profile-dir>/exports")
    args = parser.parse_args()
    profile_dir = _resolve_profile_dir("profile", args.profile_dir)
    export(profile_dir, args.out_dir)


if __name__ == "__main__":
    main()
